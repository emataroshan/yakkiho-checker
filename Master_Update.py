###########
# Master_Update.py
# Excelマスター自動更新スクリプト
###########

import openpyxl
from openpyxl import load_workbook

def update_excel_master(excel_path: str):
    # 1. ワークブックをロード（読み書き両用）
    wb = load_workbook(excel_path)

    # 2. テーブル領域を読んで辞書を作成
    #    openpyxl はテーブルオブジェクトを通じても範囲取得可能
    subcat_ws = wb['m_Subcategories']
    grp_ws    = wb['m_Groups']
    word_ws   = wb['m_Words']

    # --- 2-a. tbl_Subcategories 辞書化 ---
    # ヘッダ行を探して、列位置をマッピング
    headers = {cell.value: idx for idx, cell in enumerate(next(subcat_ws.iter_rows(min_row=1, max_row=1)), start=1)}
    scid_col   = headers['サブカテゴリID']
    scname_col = headers['サブカテゴリ名']

    subcat_map = {}
    for row in subcat_ws.iter_rows(min_row=2, values_only=True):
        scid, scname = row[scid_col-1], row[scname_col-1]
        if scid:
            subcat_map[scid] = scname

    # --- 2-b. tbl_Groups 辞書化 ---
    headers = {cell.value: idx for idx, cell in enumerate(next(grp_ws.iter_rows(min_row=1, max_row=1)), start=1)}
    gid_col     = headers['グループID']
    gpid_col    = headers['サブカテゴリID']
    gname_col   = headers['グループ名']

    group_map = {}
    for row in grp_ws.iter_rows(min_row=2, values_only=True):
        gid, scid, gname = row[gid_col-1], row[gpid_col-1], row[gname_col-1]
        if gid:
            group_map[gid] = (gname, scid)

    # 3. m_Groups シートを更新：SubcategoryName 列を追加 or 上書き
    #    すでに列があるならその位置を、なければ末尾に作る
    def ensure_column(ws, col_name):
        # ヘッダ探索
        row1 = next(ws.iter_rows(min_row=1, max_row=1))
        for cell in row1:
            if cell.value == col_name:
                return cell.column  # 既存列番号
        # 見つからなければ末尾に追加
        new_col = len(row1) + 1
        ws.cell(row=1, column=new_col, value=col_name)
        return new_col

    # m_Groups へ SubcategoryName 列を追加取得
    scname_col_in_groups = ensure_column(grp_ws, 'サブカテゴリ名')
    for row in grp_ws.iter_rows(min_row=2):
        scid = row[gpid_col-1].value
        row[scname_col_in_groups-1].value = subcat_map.get(scid, '未定義ID')

    # 4. m_Words シートを更新：GroupName + SubcategoryName
    headers = {cell.value: idx for idx, cell in enumerate(next(word_ws.iter_rows(min_row=1, max_row=1)), start=1)}
    wgid_col = headers['グループID']

    # WordName 列
    wgroupname_col = ensure_column(word_ws, 'グループ名')
    wsubcatname_col = ensure_column(word_ws, 'サブカテゴリ名')

    for row in word_ws.iter_rows(min_row=2):
        gid = row[wgid_col-1].value
        gname, scid = group_map.get(gid, ('未定義G', None))
        row[wgroupname_col-1].value   = gname
        row[wsubcatname_col-1].value  = subcat_map.get(scid,  '未定義SC')

    # 5. 上書き保存
    wb.save(excel_path)
    print(f"Updated and saved: {excel_path}")


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='Excelマスターを更新')
    parser.add_argument(
        '-f', '--file', 
        default='NGwordマスタ.xlsm', 
        help='対象Excelファイルパス'
    )
    args = parser.parse_args()
    update_excel_master(args.file)